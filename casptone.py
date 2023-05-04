import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from random import *
import time
import re

start_num = 1
finish_num = 20

count = 0
data_num = 5066

# 엑셀 파일 로드
dataset = load_workbook('data.xlsx', data_only = True)
input_data = dataset['Sheet1']
output_data = dataset['Sheet2']

# 주소 입력
driver = webdriver.Chrome()
driver.get("https://map.naver.com")
time.sleep(3)

# 검색어 입력
for i in range(start_num, finish_num + 1):
    print(f"진행률 : {i} / {data_num} / {i/data_num*100}%")

    

    
    search_box = driver.find_element(By.XPATH,'/html/body/app/layout/div[3]/div[2]/shrinkable-layout/div/app-base/search-input-box/div/div[1]/div/input')
    time.sleep(1)
    search_box.send_keys(Keys.CONTROL + "a")
    search_box.send_keys(Keys.DELETE)
    
         
   

    search_box.send_keys(input_data.cell(i,1).value, input_data.cell(i,2).value, input_data.cell(i,3).value, '카페')
    search_box.send_keys(Keys.RETURN)
    time.sleep(1)
    
  

    searchiframe = driver.find_element(By.ID, "searchIframe")
    driver.switch_to.frame(searchiframe)

    driver.find_element(By.CSS_SELECTOR, '#app-root > div > div:nth-child(1) > div > div > div > div > div > span:nth-child(2) > a').click()
    driver.find_element(By.CSS_SELECTOR, '#app-root > div > div:nth-child(1) > div > div.sgugZ.R9QDR > div > ul > li:nth-child(2) > a').click()



    # 카페 리스트 선택 (상위 10개)

    
    for j in range(1,11):
        count += 1
        
    
        driver.find_element(By.XPATH, f'//*[@id="_pcmap_list_scroll_container"]/ul/li[{j}]/div[1]/a/div/div/span[1]').click()
        cafe_name = driver.find_element(By.XPATH, f'//*[@id="_pcmap_list_scroll_container"]/ul/li[{j}]/div[1]/a/div/div/span[1]').text
        time.sleep(1)

        # 카페 세부 정보 크롤링
        driver.switch_to.default_content()
        entryiframe = driver.find_element(By.ID, "entryIframe")
        driver.switch_to.frame(entryiframe)

        try :
            cafe_address = driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div:nth-child(6) > div > div.place_section.no_margin.vKA6F > div > div > div.O8qbU.tQY7D > div > a > span.LDgIH').text
        except :
            cafe_address = "미제공"

        try :
            cafe_time = driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div:nth-child(6) > div > div.place_section.no_margin.vKA6F > div > div > div.O8qbU.pSavy > div > a > div > div > div > span > time').text
        except :
            cafe_time = "미제공"
        
        try :
            cafe_nubmer = driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div:nth-child(7) > div > div.place_section.no_margin.vKA6F > div > div > div.O8qbU.nbXkr > div > span.xlx7Q').text
        except :
            cafe_nubmer = "미제공"

        try :
            cafe_star = driver.find_element(By.CSS_SELECTOR,' #app-root > div > div > div > div.place_section.OP4V8 > div.zD5Nm.f7aZ0 > div.dAsGb > span.PXMot.LXIwF > em').text
        except :
            cafe_star = "미제공"    

        # try:
        driver.switch_to.default_content()
        driver.switch_to.frame(entryiframe)
        time.sleep(3)
       
        try:
            driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div.place_fixed_maintab > div > div > div > div > a:nth-child(5) > span').click()
            driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div.place_fixed_maintab > div > div > div > div > a:nth-child(4) > span').click()

        except:
            try:
                driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div.place_fixed_maintab > div > div > div > div > a:nth-child(4) > span').click()
                driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div.place_fixed_maintab > div > div > div > div > a:nth-child(3) > span').click()
            except:
                try:
                    driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div.place_fixed_maintab > div > div > div > div > a:nth-child(3) > span').click()
                    driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div.place_fixed_maintab > div > div > div > div > a:nth-child(2) > span').click()
                except:
                    review = "미제공"
        
        time.sleep(3)
      
        driver.execute_script("window.scrollTo(0, window.scrollY + 500);") #스크롤을 내려서 리뷰들을 봅니다. 


        driver.execute_script("window.scrollTo(0, window.scrollY + 500);") #스크롤을 내려서 리뷰들을 봅니다. 


        review_list = [] 

        for k in range(1,11):    
            try:
                review = driver.find_element(By.CSS_SELECTOR, f'#app-root > div > div > div > div:nth-child(7) > div:nth-child(3) > div.place_section.lcndr > div.place_section_content > ul > li:nth-child({k}) > div.ZZ4OK > a').text 

                
                
                #app-root > div > div > div > div:nth-child(7) > div:nth-child(3) > div.place_section.lcndr > div.place_section_content > ul > li:nth-child(1) > div.ZZ4OK > a > span

                #app-root > div > div > div > div:nth-child(7) > div:nth-child(3) > div.place_section.lcndr > div.place_section_content > ul > li:nth-child(2) > div.ZZ4OK > a
                # print(review)
                review_list.append(review)
            except:
                review_list.append("미제공")

        
        # 크롤링 데이터 출력
        output_data.cell(count, 1).value = input_data.cell(i,1).value # 시, 군 출력
        output_data.cell(count, 2).value = input_data.cell(i,2).value # 구 출력
        output_data.cell(count, 3).value = input_data.cell(i,3).value # 동 출력
        output_data.cell(count, 4).value = cafe_name # 카페 이름 출력
        output_data.cell(count, 5).value = cafe_address # 카페 주소 출력
        output_data.cell(count, 6).value = cafe_time # 카페 운영시간 출력
        output_data.cell(count, 7).value = cafe_nubmer # 카페 연락처 출력
        output_data.cell(count, 8).value = cafe_star #별점 출력 
       
        for n in range(0,9):
                try:
                    output_data.cell(count, n+9).value = review_list[n]
     
                except:
                    output_data.cell(count, n+9).value = "미제공"
            
        dataset.save("data.xlsx")
        driver.switch_to.default_content()
        driver.switch_to.frame(searchiframe)

    driver.switch_to.default_content()
    time.sleep(2)
    
    

print(f"크롤링 완료 / 찾은 카페 수 : {count}")